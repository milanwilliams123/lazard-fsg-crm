from http.server import BaseHTTPRequestHandler
import json, base64, io
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Palette ────────────────────────────────────────────────────────────────
NAVY      = "1A3A5C"
NAVY_MID  = "2E567A"
NAVY_PALE = "D6E4F0"
NAVY_XPAL = "EEF4FB"
GOLD      = "C9A84C"
GOLD_PALE = "FDF6E3"
WHITE     = "FFFFFF"
BORDER_C  = "8AAFD4"
TEXT_DARK = "1A1A1A"
TEXT_NAVY = "1A3A5C"

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(bold=False, color=TEXT_DARK, size=10, name="Arial"):
    return Font(bold=bold, color=color, size=size, name=name)

def border(weight="thin"):
    s = Side(style=weight, color=BORDER_C)
    return Border(left=s, right=s, top=s, bottom=s)

def border_outer(pos, weight="medium"):
    med = Side(style=weight, color=BORDER_C)
    none = Side(style=None)
    if pos == "left":   return Border(left=med, top=med, bottom=med, right=none)
    elif pos == "right": return Border(right=med, top=med, bottom=med, left=none)
    else:               return Border(top=med, bottom=med, left=none, right=none)

def apply_merged_border(ws, row, ncols=6):
    for ci in range(1, ncols + 1):
        pos = "left" if ci == 1 else ("right" if ci == ncols else "mid")
        ws.cell(row, ci).border = border_outer(pos)

def apply_table_outer_border(ws, top_row, bottom_row, left_col, right_col):
    med = Side(style="medium", color=NAVY)
    for r in range(top_row, bottom_row + 1):
        for c in range(left_col, right_col + 1):
            cell = ws.cell(r, c)
            ex = cell.border
            cell.border = Border(
                left=med   if c == left_col   else ex.left,
                right=med  if c == right_col  else ex.right,
                top=med    if r == top_row    else ex.top,
                bottom=med if r == bottom_row else ex.bottom,
            )

def align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def apply(cell, fnt=None, fll=None, brd=None, aln=None):
    if fnt: cell.font = fnt
    if fll: cell.fill = fll
    if brd: cell.border = brd
    if aln: cell.alignment = aln

def style_header(cell):
    apply(cell, font(bold=True, color=WHITE, size=10), fill(NAVY), border("medium"), align("center"))

def style_row(cell, even, bold_first=False):
    fll = fill(NAVY_PALE) if even else fill(NAVY_XPAL)
    apply(cell, font(bold=bold_first, color=TEXT_DARK), fll, border(), align("left", wrap=False))

def style_row_wrap(cell, even, bold_first=False):
    fll = fill(NAVY_PALE) if even else fill(NAVY_XPAL)
    apply(cell, font(bold=bold_first, color=TEXT_DARK), fll, border(), align("left", wrap=True))

def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def build_workbook(data):
    activities = data.get("activities", [])
    deals      = data.get("deals", [])
    portcos    = data.get("portcos", [])
    team       = data.get("team", [])
    selected   = data.get("selected", {})

    n_act   = len(activities)
    n_pc    = len(portcos)
    n_deals = len(deals)

    wb = Workbook()

    # ── DASHBOARD ─────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Dashboard"
    ws.sheet_view.showGridLines = False

    # Title block rows 1-3
    for r in range(1, 4):
        for ci in range(1, 7):
            apply(ws.cell(r, ci), fll=fill(NAVY))
    ws.merge_cells("A1:F1"); ws.merge_cells("A2:F2"); ws.merge_cells("A3:F3")
    ws.row_dimensions[1].height = 38
    ws.row_dimensions[2].height = 20
    ws.row_dimensions[3].height = 6

    ws["A1"].value = "LAZARD  ·  FINANCIAL SPONSORS GROUP"
    apply(ws["A1"], font(bold=True, color=WHITE, size=16), fill(NAVY), None, align("left", "center"))
    today = datetime.today().strftime("%Y-%m-%d")
    ws["A2"].value = f"FSG Coverage Intelligence  ·  Export {today}"
    apply(ws["A2"], font(bold=False, color=NAVY_PALE, size=10), fill(NAVY), None, align("left", "center"))

    # Spacer row 4
    ws.row_dimensions[4].height = 8
    for ci in range(1, 7): apply(ws.cell(4, ci), fll=fill(NAVY_XPAL))

    # KEY METRICS rows 5-7
    ws.row_dimensions[5].height = 18; ws.row_dimensions[6].height = 22; ws.row_dimensions[7].height = 44
    ws.merge_cells("A5:F5")
    ws["A5"].value = "KEY METRICS"
    apply(ws["A5"], font(bold=True, color="7A5200", size=11), fill(GOLD), None, align("left", "center"))
    apply_merged_border(ws, 5)

    kpi_labels = ["Meetings This Month", "Pitches in Flight", "Overdue Follow-ups",
                  "Tier 1 Gone Cold (60d+)", "Portcos Exploring Exit", "Pitch Win Rate"]
    kpi_formulas = [
        f"=COUNTIFS('Activity Log'!B2:B{n_act+1},\"Meeting\",'Activity Log'!A2:A{n_act+1},\">=\"&DATE(2025,3,1))",
        f"=COUNTIFS('Activity Log'!B2:B{n_act+1},\"Pitch\",'Activity Log'!M2:M{n_act+1},\"Pending\")",
        f"=COUNTIFS('Activity Log'!L2:L{n_act+1},\"<>\",'Activity Log'!M2:M{n_act+1},\"Pending\")+COUNTIFS('Activity Log'!L2:L{n_act+1},\"<>\",'Activity Log'!M2:M{n_act+1},\"In Progress\")",
        f"=COUNTIF('Activity Log'!M2:M{n_act+1},\"Pending\")",
        f"=COUNTIF('Portfolio Companies'!D2:D{n_pc+1},\"Exploring Exit\")",
        f"=IFERROR(TEXT(COUNTIF('Activity Log'!M2:M{n_act+1},\"Won\")/(COUNTIF('Activity Log'!M2:M{n_act+1},\"Won\")+COUNTIF('Activity Log'!M2:M{n_act+1},\"Lost\")),\"0%\"),\"N/A\")",
    ] if n_act > 0 else ["N/A"] * 6

    for ci, (lbl, formula) in enumerate(zip(kpi_labels, kpi_formulas), 1):
        lc = ws.cell(6, ci); lc.value = lbl
        apply(lc, font(bold=True, color=TEXT_NAVY, size=9), fill(NAVY_PALE), border("medium"), align("center", "center"))
        vc = ws.cell(7, ci); vc.value = formula
        apply(vc, font(bold=True, color=TEXT_NAVY, size=22), fill(GOLD_PALE), border("medium"), align("center", "center"))

    # Spacer row 8
    ws.row_dimensions[8].height = 10
    for ci in range(1, 7): apply(ws.cell(8, ci), fll=fill(NAVY_XPAL))

    # ACTIVITY BREAKDOWN
    ws.row_dimensions[9].height = 20
    ws.merge_cells("A9:F9")
    ws["A9"].value = "ACTIVITY BREAKDOWN"
    apply(ws["A9"], font(bold=True, color="7A5200", size=11), fill(GOLD), None, align("left", "center"))
    apply_merged_border(ws, 9)

    act_hdr = 10
    ws.row_dimensions[act_hdr].height = 20
    for ci, txt in enumerate(["Type", "Count", "", "Firm", "Touchpoints", ""], 1):
        c = ws.cell(act_hdr, ci); c.value = txt or None
        if ci in (1, 2, 4, 5):
            apply(c, font(bold=True, color=WHITE, size=10), fill(NAVY_MID), border("medium"), align("left", "center"))
        else:
            c.fill = fill(WHITE); c.border = Border()

    act_types  = ["Meeting", "Pitch", "Conference", "Call"]
    firm_names = sorted(set(a.get("firm", "") for a in activities if a.get("firm")),
                        key=lambda f: sum(1 for a in activities if a.get("firm") == f), reverse=True)[:8]
    nrows_act = max(len(act_types), len(firm_names))

    for i in range(nrows_act):
        r = act_hdr + 1 + i
        ws.row_dimensions[r].height = 18
        fll_u = fill(NAVY_PALE) if i % 2 == 0 else fill(NAVY_XPAL)
        t = act_types[i] if i < len(act_types) else ""
        firm = firm_names[i] if i < len(firm_names) else ""
        ws.cell(r, 1).value = t
        apply(ws.cell(r, 1), font(bold=bool(t), color=TEXT_DARK), fll_u, border(), align("left", "center"))
        if t:
            ws.cell(r, 2).value = f"=COUNTIF('Activity Log'!B2:B{n_act+1},\"{t}\")"
        apply(ws.cell(r, 2), font(color=TEXT_DARK), fll_u, border(), align("center", "center"))
        ws.cell(r, 3).value = None; ws.cell(r, 3).fill = fill(WHITE); ws.cell(r, 3).border = Border()
        ws.cell(r, 4).value = firm
        apply(ws.cell(r, 4), font(bold=bool(firm), color=TEXT_DARK), fll_u, border(), align("left", "center"))
        if firm:
            ws.cell(r, 5).value = f"=COUNTIF('Activity Log'!C2:C{n_act+1},\"{firm}\")"
        apply(ws.cell(r, 5), font(color=TEXT_DARK), fll_u, border(), align("center", "center"))
        ws.cell(r, 6).value = None; ws.cell(r, 6).fill = fill(WHITE); ws.cell(r, 6).border = Border()

    act_last = act_hdr + nrows_act
    apply_table_outer_border(ws, act_hdr, act_last, 1, 2)
    apply_table_outer_border(ws, act_hdr, act_last, 4, 5)

    # DEAL PIPELINE
    spc_r = act_hdr + 1 + nrows_act
    ws.row_dimensions[spc_r].height = 8
    for ci in range(1, 7): apply(ws.cell(spc_r, ci), fll=fill(NAVY_XPAL))

    deal_sec = spc_r + 1
    ws.row_dimensions[deal_sec].height = 20
    ws.merge_cells(f"A{deal_sec}:F{deal_sec}")
    ws.cell(deal_sec, 1).value = "DEAL PIPELINE"
    apply(ws.cell(deal_sec, 1), font(bold=True, color="7A5200", size=11), fill(GOLD), None, align("left", "center"))
    apply_merged_border(ws, deal_sec)

    deal_hdr = deal_sec + 1
    ws.row_dimensions[deal_hdr].height = 20
    for ci, txt in enumerate(["Deal Name", "Role", "Status", "Seller", "Asset", "Date"], 1):
        c = ws.cell(deal_hdr, ci); c.value = txt
        apply(c, font(bold=True, color=WHITE, size=10), fill(NAVY_MID), border("medium"), align("left", "center"))

    for i in range(n_deals):
        r = deal_hdr + 1 + i
        ws.row_dimensions[r].height = 18
        fll_u = fill(NAVY_PALE) if i % 2 == 0 else fill(NAVY_XPAL)
        dr = i + 2
        for ci, formula in enumerate([f"=Deals!A{dr}", f"=Deals!C{dr}", f"=Deals!D{dr}",
                                       f"=Deals!F{dr}", f"=Deals!G{dr}", f"=Deals!B{dr}"], 1):
            c = ws.cell(r, ci); c.value = formula
            apply(c, font(bold=(ci == 1), color=TEXT_DARK), fll_u, border(), align("left", "center"))

    # PORTCOS EXPLORING EXIT
    exit_pcs = [(i, p) for i, p in enumerate(portcos) if p.get("status") == "Exploring Exit"]
    spc_r2 = deal_hdr + 1 + n_deals
    ws.row_dimensions[spc_r2].height = 8
    for ci in range(1, 7): apply(ws.cell(spc_r2, ci), fll=fill(NAVY_XPAL))

    exit_sec = spc_r2 + 1
    ws.row_dimensions[exit_sec].height = 20
    ws.merge_cells(f"A{exit_sec}:F{exit_sec}")
    ws.cell(exit_sec, 1).value = "PORTCOS EXPLORING EXIT"
    apply(ws.cell(exit_sec, 1), font(bold=True, color="7A5200", size=11), fill(GOLD), None, align("left", "center"))
    apply_merged_border(ws, exit_sec)

    exit_hdr = exit_sec + 1
    ws.row_dimensions[exit_hdr].height = 20
    for ci, txt in enumerate(["Company", "Firm", "Sector", "Revenue", "EBITDA", "PE Owner"], 1):
        c = ws.cell(exit_hdr, ci); c.value = txt
        apply(c, font(bold=True, color=WHITE, size=10), fill(NAVY_MID), border("medium"), align("left", "center"))

    for j, (orig_i, _) in enumerate(exit_pcs):
        r = exit_hdr + 1 + j
        ws.row_dimensions[r].height = 18
        fll_u = fill(NAVY_PALE) if j % 2 == 0 else fill(NAVY_XPAL)
        pc_row = orig_i + 2
        for ci, formula in enumerate([f"='Portfolio Companies'!A{pc_row}", f"='Portfolio Companies'!B{pc_row}",
                                       f"='Portfolio Companies'!C{pc_row}", f"='Portfolio Companies'!E{pc_row}",
                                       f"='Portfolio Companies'!F{pc_row}", f"='Portfolio Companies'!H{pc_row}"], 1):
            c = ws.cell(r, ci); c.value = formula
            apply(c, font(bold=(ci == 1), color=TEXT_DARK), fll_u, border(), align("left", "center"))

    set_col_widths(ws, [32, 16, 18, 26, 24, 14])
    wb.calculation.calcMode = "auto"
    wb.calculation.fullCalcOnLoad = True

    # ── ACTIVITY LOG ──────────────────────────────────────────────────────
    if selected.get("activity", True) and activities:
        ws2 = wb.create_sheet("Activity Log")
        ws2.sheet_view.showGridLines = False
        cols = ["Date","Type","Firm","Contact","Lazard Banker","Deal Side","Linked Deal",
                "Portfolio Company","Description","Notes","Outcome","Next Action","Status"]
        widths = [12,13,18,20,20,12,24,22,34,40,26,30,13]
        ws2.row_dimensions[1].height = 24
        for ci, h in enumerate(cols, 1):
            style_header(ws2.cell(1, ci)); ws2.cell(1, ci).value = h
        for ri, a in enumerate(activities, 2):
            ws2.row_dimensions[ri].height = 18
            vals = [a.get("date",""),a.get("type",""),a.get("firm",""),a.get("contact",""),
                    a.get("banker",""),a.get("dealSide",""),a.get("deal",""),a.get("portco",""),
                    a.get("description",""),a.get("notes",""),a.get("outcome",""),
                    a.get("nextAction",""),a.get("status","")]
            even = ri % 2 == 0
            for ci, val in enumerate(vals, 1):
                c = ws2.cell(ri, ci); c.value = val
                fn = style_row_wrap if ci in (9, 10, 11, 12) else style_row
                fn(c, even, bold_first=(ci == 1))
        set_col_widths(ws2, widths)
        ws2.freeze_panes = "A2"

    # ── DEALS ─────────────────────────────────────────────────────────────
    if selected.get("deals", True) and deals:
        ws3 = wb.create_sheet("Deals")
        ws3.sheet_view.showGridLines = False
        cols = ["Deal Name","Date","Lazard Role","Status","Sector","Seller Firm",
                "Portfolio Company","Buy-Side Firms Pitched","Description","Notes"]
        widths = [30,13,14,14,18,18,22,34,38,38]
        ws3.row_dimensions[1].height = 24
        for ci, h in enumerate(cols, 1):
            style_header(ws3.cell(1, ci)); ws3.cell(1, ci).value = h
        for ri, d in enumerate(deals, 2):
            ws3.row_dimensions[ri].height = 18
            vals = [d.get("name",""),d.get("date",""),d.get("role",""),d.get("status",""),
                    d.get("sector",""),d.get("seller",""),d.get("portco",""),d.get("buyside",""),
                    d.get("description",""),d.get("notes","")]
            even = ri % 2 == 0
            for ci, val in enumerate(vals, 1):
                c = ws3.cell(ri, ci); c.value = val
                fn = style_row_wrap if ci >= 9 else style_row
                fn(c, even, bold_first=(ci == 1))
        set_col_widths(ws3, widths)
        ws3.freeze_panes = "A2"

    # ── PORTFOLIO COMPANIES ────────────────────────────────────────────────
    if selected.get("portcos", True) and portcos:
        ws4 = wb.create_sheet("Portfolio Companies")
        ws4.sheet_view.showGridLines = False
        cols = ["Company","Firm","Sector","Status","Revenue","EBITDA","Investment Year","PE Owner Contact","Lazard Notes"]
        widths = [24,16,18,16,12,12,16,22,42]
        ws4.row_dimensions[1].height = 24
        for ci, h in enumerate(cols, 1):
            style_header(ws4.cell(1, ci)); ws4.cell(1, ci).value = h
        for ri, p in enumerate(portcos, 2):
            ws4.row_dimensions[ri].height = 18
            vals = [p.get("name",""),p.get("firm",""),p.get("sector",""),p.get("status",""),
                    p.get("revenue",""),p.get("ebitda",""),p.get("year",""),p.get("owner",""),p.get("notes","")]
            even = ri % 2 == 0
            for ci, val in enumerate(vals, 1):
                c = ws4.cell(ri, ci); c.value = val
                fn = style_row_wrap if ci == 9 else style_row
                fn(c, even, bold_first=(ci == 1))
        set_col_widths(ws4, widths)
        ws4.freeze_panes = "A2"

    # ── LAZARD TEAM ───────────────────────────────────────────────────────
    if selected.get("team", True) and team:
        ws5 = wb.create_sheet("Lazard Team")
        ws5.sheet_view.showGridLines = False
        cols = ["Name","Title","Role / Sector","Email","Phone","PE Coverage","Activities YTD"]
        widths = [22,24,24,28,18,52,14]
        ws5.row_dimensions[1].height = 24
        for ci, h in enumerate(cols, 1):
            style_header(ws5.cell(1, ci)); ws5.cell(1, ci).value = h
        for ri, b in enumerate(team, 2):
            ws5.row_dimensions[ri].height = 22
            vals = [b.get("name",""),b.get("title",""),b.get("role",""),b.get("email",""),
                    b.get("phone",""),b.get("coverage",""),b.get("activitiesYTD", 0)]
            even = ri % 2 == 0
            for ci, val in enumerate(vals, 1):
                c = ws5.cell(ri, ci); c.value = val
                fn = style_row_wrap if ci == 6 else style_row
                fn(c, even, bold_first=(ci == 1))
        set_col_widths(ws5, widths)
        ws5.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


class handler(BaseHTTPRequestHandler):
    def do_OPTIONS(self):
        self.send_response(200)
        self.send_header("Access-Control-Allow-Origin", "*")
        self.send_header("Access-Control-Allow-Methods", "POST, OPTIONS")
        self.send_header("Access-Control-Allow-Headers", "Content-Type")
        self.end_headers()

    def do_POST(self):
        length = int(self.headers.get("Content-Length", 0))
        body = self.rfile.read(length)
        try:
            data = json.loads(body)
            file_bytes = build_workbook(data)
            today = datetime.today().strftime("%Y-%m-%d")
            filename = f"Lazard_FSG_Export_{today}.xlsx"
            self.send_response(200)
            self.send_header("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            self.send_header("Content-Disposition", f'attachment; filename="{filename}"')
            self.send_header("Content-Length", str(len(file_bytes)))
            self.send_header("Access-Control-Allow-Origin", "*")
            self.end_headers()
            self.wfile.write(file_bytes)
        except Exception as e:
            self.send_response(500)
            self.send_header("Content-Type", "application/json")
            self.send_header("Access-Control-Allow-Origin", "*")
            self.end_headers()
            self.wfile.write(json.dumps({"error": str(e)}).encode())